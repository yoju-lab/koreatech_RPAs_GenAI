import os
import sys
import json
import shutil
import datetime
import urllib.request
import pandas as pd
import openpyxl
from openai import OpenAI
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# API credentials
client_id = os.getenv("NAVER_CLIENT_ID")
client_secret = os.getenv("NAVER_CLIENT_SECRET")
openai_api_key = os.getenv("OPENAI_API_KEY")
encText = urllib.parse.quote("포켄스")

# OpenAI client setup
client = OpenAI(api_key=openai_api_key)
openai_model = "gpt-4o-mini"

# File paths
current_folder = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(current_folder, 'genai_rpa.xlsx')


def create_workbook_if_not_exists():
    """
    Ensure the workbook exists. If not, create it with default sheets.
    """
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.active.title = "now_list"
        wb.create_sheet(title="now_report")
        wb.save(file_path)
        print(f"Workbook created: {file_path}")
    else:
        print(f"Workbook already exists: {file_path}")


def convert_json_to_dataframe(json_result):
    """
    Convert JSON result to a pandas DataFrame with an added '순위' column.
    """
    if isinstance(json_result, str):
        json_result = json.loads(json_result)
    items = json_result.get('items', [])
    df = pd.DataFrame(items)
    df.insert(0, "순위", range(1, len(df) + 1))
    df.set_index("순위", inplace=True)
    return df


def fetch_naver_api_data(api_type):
    """
    Fetch data from Naver API (shopping or news) based on the given type.
    """
    base_url = f"https://openapi.naver.com/v1/search/{api_type}?sort=date&display=20&query={encText}"
    request = urllib.request.Request(base_url)
    request.add_header("X-Naver-Client-Id", client_id)
    request.add_header("X-Naver-Client-Secret", client_secret)
    response = urllib.request.urlopen(request)
    if response.getcode() == 200:
        return response.read().decode('utf-8')
    else:
        print(f"Error Code: {response.getcode()}")
        return None


def handle_list_sheet(wb):
    """
    Manage 'now_list' and 'prev_list' sheets in the workbook.
    """
    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    backup_file = os.path.join(current_folder, f"genai_rpa_{timestamp}.xlsx")
    shutil.copy(file_path, backup_file)
    print(f"Backup created: {backup_file}")

    if 'prev_list' in wb.sheetnames:
        del wb['prev_list']
        print("Sheet 'prev_list' deleted.")

    if 'now_list' in wb.sheetnames:
        wb['now_list'].title = 'prev_list'
        print("Sheet 'now_list' renamed to 'prev_list'.")
    else:
        print("Sheet 'now_list' not found.")
        return False

    wb.create_sheet(title="now_list", index=wb.sheetnames.index('prev_list') + 1)
    print("Sheet 'now_list' created.")
    return True


def update_sheet_with_dataframe(sheet, dataframe):
    """
    Clear and update the given sheet with data from the DataFrame.
    """
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    print(f"Sheet '{sheet.title}' updated.")


def call_openai_api(prompt):
    """
    Call OpenAI API with the given prompt and return the response.
    """
    completion = client.chat.completions.create(
        model=openai_model,
        messages=[{"role": "user", "content": prompt}]
    )
    return completion.choices[0].message.content


def generate_analysis_prompt(prev_data, now_data):
    """
    Generate a prompt for analyzing shopping list changes.
    """
    return f"""
    너는 데이터분석 전문가야.
    다음 두 상품 목록을 비교 분석해 변화 패턴을 도출해주세요:
    
    prev_list(변경 전): {prev_data}
    now_list(변경 후): {now_data}
    
    분석 요구사항:
    1. 상품 정보의 구조적 변화(형식, 필드값 등) 파악
    2. 상품 가격, 재고, 카테고리 등 주요 속성 변화 탐지
    3. 쇼핑몰별 상품 분포 변화 분석
    4. 브랜드/제조사 정보 변경 사항 확인
    5. 새로 추가되거나 삭제된 상품 식별
    
    결과물 요청사항:
    - 변화의 핵심 패턴을 3-5개 포인트로 요약
    - 구체적인 수치, 상품명, 쇼핑몰명을 포함하여 근거 제시
    - 한글로 작성, 총 400-500자 이내로 간결하게 작성
    - 마크다운, HTML 태그, 특수기호 사용 금지
    - 실제 소비자에게 유용한 인사이트 중심으로 작성
    """


def update_report_sheet(sheet, title, content, start_row):
    """
    Update the report sheet with the given title and content starting from a specific row.
    """
    sheet.cell(row=start_row, column=1, value=title)
    sheet.cell(row=start_row + 1, column=1, value=content)
    sheet.cell(row=start_row + 1, column=1).alignment = openpyxl.styles.Alignment(wrap_text=True)
    print(f"Report updated with '{title}'.")


def main():
    create_workbook_if_not_exists()
    wb = openpyxl.load_workbook(file_path)

    if not handle_list_sheet(wb):
        wb.close()
        return

    shopping_data = fetch_naver_api_data("shop")
    if shopping_data:
        df_shopping = convert_json_to_dataframe(shopping_data)
        update_sheet_with_dataframe(wb['now_list'], df_shopping)

        prev_data = [[cell.value for cell in row] for row in wb['prev_list'].iter_rows()]
        now_data = [[cell.value for cell in row] for row in wb['now_list'].iter_rows()]
        analysis_prompt = generate_analysis_prompt(prev_data, now_data)
        analysis_result = call_openai_api(analysis_prompt)
        update_report_sheet(wb['now_report'], "오픈 마켓 리포트", analysis_result, 4)

    news_data = fetch_naver_api_data("news")
    if news_data:
        news_prompt = f"""
        너는 뉴스 요약 전문가야.
        다음 뉴스 내용을 요약해주세요:
        
        뉴스 내용: {news_data}
        
        요약 요구사항:
        1. 주요 뉴스 주제 및 핵심 메시지 요약
        2. 구체적인 수치, 고유명사, 키워드 포함
        3. 소비자에게 유용한 인사이트 제공
        
        결과물 요청사항:
        - 한글로 작성, 총 300-400자 이내로 간결하게 작성
        - 글머리를 활용하여 명확하고 간결한 요약 작성
        - 마크다운, HTML 태그, 특수기호 사용 금지
        """
        news_summary = call_openai_api(news_prompt)
        update_report_sheet(wb['now_report'], "네이버 뉴스 분석", news_summary, 7)

    wb.save(file_path)
    print("Workbook saved and closed.")


if __name__ == '__main__':
    main()
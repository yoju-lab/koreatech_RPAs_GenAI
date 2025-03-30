import openpyxl
import datetime
import shutil
import pandas as pd
import json
import os
import sys
import urllib.request
from openai import OpenAI

import os
from dotenv import load_dotenv

# .env 파일 로드
load_dotenv()

# .env에서 값 가져오기
client_id = os.getenv("NAVER_CLIENT_ID")
client_secret = os.getenv("NAVER_CLIENT_SECRET")
encText = urllib.parse.quote("포켄스")

# OpenAI API 호출
client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
)
openai_model = "gpt-4o-mini"

# 1. genai_rpa.xlsx 파일 경로 설정
current_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(current_folder, 'genai_rpa.xlsx')


def create_workbook_if_not_exists():
    """
    Check if the genai_rpa.xlsx file exists. If not, create it with the required sheets.
    """
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        # Create default sheets
        wb.active.title = "now_list"
        wb.create_sheet(title="now_report")
        wb.save(file_path)
        print(f"Workbook created: {file_path}")
    else:
        print(f"Workbook already exists: {file_path}")


def convert_json_to_dataframe(json_result):
    """
    json_result가 문자열이면 dict로 변환한 후,
    "items" 데이터를 pandas DataFrame으로 변환하고,
    "순위" 열을 제일 왼쪽에 추가하여 1부터 일련번호를 부여한 후
    이 열을 index로 설정합니다.
    """
    if isinstance(json_result, str):
        json_result = json.loads(json_result)
    items = json_result.get('items', [])
    df = pd.DataFrame(items)
    # "순위" 열을 추가 (1부터 시작하는 일련번호)
    df.insert(0, "순위", range(1, len(df)+1))
    # "순위" 열을 index로 설정
    df.set_index("순위", inplace=True)
    return df


def get_naver_shopping_list_data():
    # display를 20으로 수정. 페이징을 한다면 start=번호 형태를 추가
    url = "https://openapi.naver.com/v1/search/shop?sort=date&display=20&query=" + encText # JSON 결과

    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if(rescode==200):
        response_body = response.read()
        # print(response_body.decode('utf-8'))
    else:
        print("Error Code:" + rescode)
    result = response_body.decode('utf-8')

    return result


def get_naver_news_data():
    """
    Fetch news data from Naver News API.
    """
    url = "https://openapi.naver.com/v1/search/news?sort=date&display=20&query=" + encText  # JSON 결과

    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id", client_id)
    request.add_header("X-Naver-Client-Secret", client_secret)
    response = urllib.request.urlopen(request)
    rescode = response.getcode()
    if rescode == 200:
        response_body = response.read()
    else:
        print("Error Code:" + rescode)
        return None

    result = response_body.decode('utf-8')
    return result


def handle_list_sheet(wb):
    # 2. 현재 시각을 기반으로 백업 파일 생성
    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    backup_file = os.path.join(current_folder, f"genai_rpa_{timestamp}.xlsx")
    shutil.copy(file_path, backup_file)
    print(f"Backup created: {backup_file}")

    # 3. 'prev_list' 시트 삭제 (존재하는 경우)
    if 'prev_list' in wb.sheetnames:
        del wb['prev_list']
        print("Sheet 'prev_list' deleted.")
    else:
        print("Sheet 'prev_list' does not exist. Skipping deletion.")

    # 4. 'now_list' 시트의 이름을 'prev_list'로 변경
    if 'now_list' in wb.sheetnames:
        now_list_sheet = wb['now_list']
        now_list_sheet.title = 'prev_list'
        print("Sheet 'now_list' renamed to 'prev_list'.")
    else:
        print("Sheet 'now_list' not found. Please check the workbook.")
        return False

    # 5. 새로운 'now_list' 시트 생성
    wb.create_sheet(title="now_list", index=wb.sheetnames.index('prev_list') + 1)
    print("Sheet 'now_list' created.")

    return True


def update_now_list(wb, df_shopping):
    # 8. 'now_list' 시트 내용 업데이트
    now_sheet = wb['now_list']
    for row in now_sheet.iter_rows():
        for cell in row:
            cell.value = None  # Clear existing content

    for r_idx, row in enumerate(df_shopping.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            now_sheet.cell(row=r_idx, column=c_idx, value=value)

    print("Sheet 'now_list' updated with new shopping data.")
    return


def conn_openai_api(prompt):
    completion = client.chat.completions.create(
        model=openai_model,
        messages=[
            {
                "role": "user",
                "content": prompt
            }
        ]
    )

    # 분석글 출력
    print(completion.choices[0].message.content)
    result = completion.choices[0].message.content

    return result


def get_openai_shopping_list_anaysis(wb):
    # 각 시트의 전체 데이터를 불러오기
    prev_sheet = wb['prev_list']
    now_sheet = wb['now_list']

    prev_data = [[cell.value for cell in row] for row in prev_sheet.iter_rows()]
    now_data = [[cell.value for cell in row] for row in now_sheet.iter_rows()]

    # 비교 분석을 위한 프롬프트 구성
    # prompt = f"""
    # 다음 두 목록을 비교 분석하여, prev_list 목록에서 now_list 목록으로 바뀐 주요 특징을 도출해줘.
    # prev_list 목록: {prev_data}
    # now_list 목록: {now_data}

    # 비교 분석 결과를 바탕으로 구체적인 수치, 상품명, 쇼핑몰명 등을 언급하며 한글로 500자 이내의 분석글을 작성해줘.
    # markdown 언어나, html 태그와 html 특수기호 등을 사용하지 말아줘.
    # """

    # 개선
    prompt = f"""
    너는 데이터분석 전문가야.
    다음 두 상품 목록을 비교 분석해 변화 패턴을 도출해주세요:
    
    prev_list(변경 전): {prev_data}
    now_list(변경 후): {now_data}
    
    분석 요구사항:
    1. 상품 정보의 구조적 변화(형식, 필드값 등) 파악
    2. 상품 가격, 재고, 카테고리 등 주요 속성 변화 탐지
    3. 쇼핑몰별 상품 분포 변화 분석
    4. 브랜드/제조사 정보 변경 사항 확인
    5. 새로 추가되거나 삭제된 상품 식별
    
    결과물 요청사항:
    - 변화의 핵심 패턴을 3-5개 포인트로 요약
    - 구체적인 수치, 상품명, 쇼핑몰명을 포함하여 근거 제시
    - 한글로 작성, 총 400-500자 이내로 간결하게 작성
    - 마크다운, HTML 태그, 특수기호 사용 금지
    - 실제 소비자에게 유용한 인사이트 중심으로 작성
    """
    result = conn_openai_api(prompt)

    return result


def get_openai_news_summarize(result_news):
    """
    Summarize the news content using OpenAI API.
    """
    prompt = f"""
    너는 뉴스 요약 전문가야.
    다음 뉴스 내용을 요약해주세요:
    
    뉴스 내용: {result_news}
    
    요약 요구사항:
    1. 주요 뉴스 주제 및 핵심 메시지 요약
    2. 구체적인 수치, 고유명사, 키워드 포함
    3. 소비자에게 유용한 인사이트 제공
    
    결과물 요청사항:
    - 한글로 작성, 총 300-400자 이내로 간결하게 작성
    - 글머리를 활용하여 명확하고 간결한 요약 작성
    - 마크다운, HTML 태그, 특수기호 사용 금지
    """
    result = conn_openai_api(prompt)
    return result


def update_now_report(wb, analysis):
    now_sheet = wb['now_report']

    # 현재 날짜와 시간(년-월-일 시:분:초) 포맷팅
    current_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    now_sheet.cell(row=3, column=1, value=f"{current_dt} 기준")

    now_sheet.cell(row=4, column=1, value=f"오픈 마켓 리포트")

    now_sheet.cell(row=5, column=1, value=analysis)
    now_sheet.cell(row=5, column=1).alignment = openpyxl.styles.Alignment(wrap_text=True)
    print("Sheet 'now_report' updated with new analysis.")

    return


def update_now_report_with_news(wb, news_analysis):
    """
    Add Naver News analysis to the 'now_report' sheet.
    """
    now_sheet = wb['now_report']

    # Add news analysis below the shopping analysis
    now_sheet.cell(row=7, column=1, value="네이버 뉴스 분석")
    now_sheet.cell(row=8, column=1, value=news_analysis)
    now_sheet.cell(row=8, column=1).alignment = openpyxl.styles.Alignment(wrap_text=True)
    print("Sheet 'now_report' updated with Naver News analysis.")

    return


def save_close_file(wb):
    # 9. genai_rpa.xlsx 파일 저장
    wb.save(file_path)
    print("Workbook saved and closed.")
    return


def main():
    # Ensure the workbook exists
    create_workbook_if_not_exists()

    # Open the workbook
    wb = openpyxl.load_workbook(file_path)

    # 2~5
    if not handle_list_sheet(wb):
        wb.close()
        return

    # 6. 네이버 API 함수를 호출하여 쇼핑 목록 데이터 가져오기
    result_json = get_naver_shopping_list_data()

    # 7. result JSON을 pandas DataFrame 형태로 만들기 (순위 열 추가됨)
    df_shopping = convert_json_to_dataframe(result_json)
    print("Converted JSON to DataFrame.")

    # 8. 'now_list' 시트 내용 업데이트
    update_now_list(wb, df_shopping)

    # 9. OpenAI 분석 결과 생성
    result_analysis = get_openai_shopping_list_anaysis(wb)

    # 10. 분석 결과 업데이트
    update_now_report(wb, result_analysis)

    # 12. 네이버 뉴스 데이터 가져오기
    result_news_json = get_naver_news_data()

    if result_news_json:
        # 13. OpenAI를 사용하여 뉴스 데이터 요약
        news_analysis = get_openai_news_summarize(result_news_json)

        # 14. 뉴스 분석 결과 업데이트
        update_now_report_with_news(wb, news_analysis)

    # 15. 파일 저장 및 종료
    save_close_file(wb)


if __name__ == '__main__':
    main()
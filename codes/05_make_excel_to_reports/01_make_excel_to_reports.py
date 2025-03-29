import os
import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook
import shutil
import requests
import urllib.request
from dotenv import load_dotenv

def get_naver_shopping_data():
    """
    네이버 쇼핑 API를 호출하여 쇼핑 목록 데이터를 가져옵니다.
    """
    # .env 파일 로드
    load_dotenv()

    # .env에서 값 가져오기
    client_id = os.getenv("NAVER_CLIENT_ID")
    client_secret = os.getenv("NAVER_CLIENT_SECRET")

    if not client_id or not client_secret:
        raise ValueError("NAVER_CLIENT_ID 또는 NAVER_CLIENT_SECRET 환경 변수가 설정되지 않았습니다.")

    encText = urllib.parse.quote("쇼핑")
    url = f"https://openapi.naver.com/v1/search/shop?display=20&sort=date&query={encText}"
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id", client_id)
    request.add_header("X-Naver-Client-Secret", client_secret)

    response = urllib.request.urlopen(request)
    rescode = response.getcode()

    if rescode == 200:
        response_body = response.read().decode('utf-8')  # Decode bytes to string
        # JSON 데이터를 Pandas DataFrame으로 변환
        shopping_data = pd.read_json(response_body)
        return shopping_data
    else:
        raise Exception(f"네이버 쇼핑 API 호출 실패. Error Code: {rescode}")

def main():
    # 현재 작업 폴더 경로
    current_folder = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(current_folder, "genai_rpa.xlsx")
    
    # 1. 현재 폴더에 genai_rpa.xlsx 없으면 생성
    if not os.path.exists(excel_file):
        wb = Workbook()
        # 기본 시트 이름 변경
        ws = wb.active
        ws.title = "now_list"
        # prev_list 시트 추가
        wb.create_sheet("prev_list")
        wb.save(excel_file)
        print(f"파일 생성: {excel_file}")
    
    # 2. genai_rpa_yyyymmddHHMMSS.xlsx 백업
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y%m%d%H%M%S")
    backup_file = os.path.join(current_folder, f"genai_rpa_{timestamp}.xlsx")
    
    if os.path.exists(excel_file):
        shutil.copy2(excel_file, backup_file)
        print(f"백업 파일 생성: {backup_file}")
    
    # 3, 4, 5. 시트 관리 (시트 제거, 이름 변경, 생성)
    wb = openpyxl.load_workbook(excel_file)
    
    # 3. prev_list 시트가 있으면 제거
    if "prev_list" in wb.sheetnames:
        del wb["prev_list"]
        print("prev_list 시트 제거 완료")
    
    # 4. now_list 시트를 prev_list로 변경
    if "now_list" in wb.sheetnames:
        now_list_sheet = wb["now_list"]
        now_list_sheet.title = "prev_list"
        print("now_list 시트를 prev_list로 변경 완료")
    
    # 5. now_list 시트 생성
    wb.create_sheet("now_list")
    print("now_list 시트 생성 완료")
    
    # 일단 변경사항 저장
    wb.save(excel_file)
    
    # 6. naver api function 호출 : 쇼핑목록 데이터 가져오기
    shopping_data = get_naver_shopping_data()
    print("네이버 쇼핑 데이터 가져오기 완료")
    
    # 7. now_list sheet 내용 update
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        shopping_data.to_excel(writer, sheet_name="now_list", index=False)
    print("now_list 시트 데이터 업데이트 완료")
    
    # 8. genai_rpa.xlsx 파일 저장 (이미 위에서 저장되었음)
    print(f"작업이 완료되었습니다. 파일 경로: {excel_file}")

if __name__ == "__main__":
    main()
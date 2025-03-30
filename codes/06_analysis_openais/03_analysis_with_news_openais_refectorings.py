import openpyxl
import datetime
import shutil
import pandas as pd
import json
import os
import urllib.request
from dotenv import load_dotenv
from openai import OpenAI

# Load environment variables
load_dotenv()

# Retrieve API keys and constants
client_id = os.getenv("NAVER_CLIENT_ID")
client_secret = os.getenv("NAVER_CLIENT_SECRET")
openai_api_key = os.getenv("OPENAI_API_KEY")
encText = urllib.parse.quote("포켄스")
current_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(current_folder, 'genai_rpa.xlsx')

# Initialize OpenAI client
client = OpenAI(api_key=openai_api_key)
openai_model = "gpt-4o-mini"


def create_workbook_if_not_exists():
    """
    Ensure the workbook exists. If not, create it with default sheets.
    """
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.active.title = "now_list"
        wb.create_sheet(title="now_report")
        wb.save(file_path)
        print(f"Workbook created: {file_path}")
    else:
        print(f"Workbook already exists: {file_path}")


def convert_json_to_dataframe(json_result):
    """
    Convert JSON result to a pandas DataFrame with a ranking column.
    """
    if isinstance(json_result, str):
        json_result = json.loads(json_result)
    items = json_result.get('items', [])
    df = pd.DataFrame(items)
    df.insert(0, "순위", range(1, len(df) + 1))
    df.set_index("순위", inplace=True)
    return df


def fetch_naver_api_data(api_url):
    """
    Fetch data from the Naver API.
    """
    request = urllib.request.Request(api_url)
    request.add_header("X-Naver-Client-Id", client_id)
    request.add_header("X-Naver-Client-Secret", client_secret)
    response = urllib.request.urlopen(request)
    if response.getcode() == 200:
        return response.read().decode('utf-8')
    else:
        print(f"Error Code: {response.getcode()}")
        return None


def handle_list_sheet(wb):
    """
    Manage the 'now_list' and 'prev_list' sheets in the workbook.
    """
    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    backup_file = os.path.join(current_folder, f"genai_rpa_{timestamp}.xlsx")
    shutil.copy(file_path, backup_file)
    print(f"Backup created: {backup_file}")

    if 'prev_list' in wb.sheetnames:
        del wb['prev_list']
        print("Sheet 'prev_list' deleted.")

    if 'now_list' in wb.sheetnames:
        wb['now_list'].title = 'prev_list'
        print("Sheet 'now_list' renamed to 'prev_list'.")
    else:
        print("Sheet 'now_list' not found. Please check the workbook.")
        return False

    wb.create_sheet(title="now_list", index=wb.sheetnames.index('prev_list') + 1)
    print("Sheet 'now_list' created.")
    return True


def update_sheet_with_dataframe(sheet, dataframe):
    """
    Update a worksheet with the contents of a DataFrame.
    """
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)


def generate_openai_analysis(prompt):
    """
    Generate analysis using OpenAI API based on the provided prompt.
    """
    completion = client.chat.completions.create(
        model=openai_model,
        messages=[{"role": "user", "content": prompt}]
    )
    return completion.choices[0].message.content


def update_report_sheet(sheet, analysis, row_start):
    """
    Update the report sheet with analysis content.
    """
    current_dt = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.cell(row=row_start, column=1, value=f"{current_dt} 기준")
    sheet.cell(row=row_start + 1, column=1, value=analysis)
    sheet.cell(row=row_start + 1, column=1).alignment = openpyxl.styles.Alignment(wrap_text=True)


def main():
    create_workbook_if_not_exists()
    wb = openpyxl.load_workbook(file_path)

    if not handle_list_sheet(wb):
        wb.close()
        return

    shopping_data = fetch_naver_api_data(
        f"https://openapi.naver.com/v1/search/shop?sort=date&display=20&query={encText}"
    )
    if shopping_data:
        df_shopping = convert_json_to_dataframe(shopping_data)
        update_sheet_with_dataframe(wb['now_list'], df_shopping)

        analysis_prompt = f"""
        너는 데이터분석 전문가야.
        다음 두 상품 목록을 비교 분석해 변화 패턴을 도출해주세요:
        
        prev_list(변경 전): {[[cell.value for cell in row] for row in wb['prev_list'].iter_rows()]}
        now_list(변경 후): {[[cell.value for cell in row] for row in wb['now_list'].iter_rows()]}
        
        분석 요구사항:
        1. 상품 정보의 구조적 변화(형식, 필드값 등) 파악
        2. 상품 가격, 재고, 카테고리 등 주요 속성 변화 탐지
        3. 쇼핑몰별 상품 분포 변화 분석
        4. 브랜드/제조사 정보 변경 사항 확인
        5. 새로 추가되거나 삭제된 상품 식별
        
        결과물 요청사항:
        - 변화의 핵심 패턴을 3-5개 포인트로 요약
        - 구체적인 수치, 상품명, 쇼핑몰명을 포함하여 근거 제시
        - 한글로 작성, 총 400-500자 이내로 간결하게 작성
        - 마크다운, HTML 태그, 특수기호 사용 금지
        """
        analysis = generate_openai_analysis(analysis_prompt)
        update_report_sheet(wb['now_report'], analysis, row_start=3)

    news_data = fetch_naver_api_data(
        f"https://openapi.naver.com/v1/search/news?sort=date&display=20&query={encText}"
    )
    if news_data:
        news_prompt = f"""
        너는 뉴스 요약 전문가야.
        다음 뉴스 내용을 요약해주세요:
        
        뉴스 내용: {news_data}
        
        요약 요구사항:
        1. 주요 뉴스 주제 및 핵심 메시지 요약
        2. 구체적인 수치, 고유명사, 키워드 포함
        3. 소비자에게 유용한 인사이트 제공
        
        결과물 요청사항:
        - 한글로 작성, 총 300-400자 이내로 간결하게 작성
        - 글머리를 활용하여 명확하고 간결한 요약 작성
        - 마크다운, HTML 태그, 특수기호 사용 금지
        """
        news_analysis = generate_openai_analysis(news_prompt)
        update_report_sheet(wb['now_report'], news_analysis, row_start=7)

    wb.save(file_path)
    print("Workbook saved and closed.")


if __name__ == '__main__':
    main()
import os
import json
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

# 환경 변수 로드
load_dotenv()

# OpenAI 클라이언트 초기화
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

def generate_curriculum(topic, description, total_hours):
    """OpenAI를 사용하여 강의 커리큘럼 생성"""
    prompt = f"""
    다음 주제와 내용에 대한 {total_hours}시간 강의 커리큘럼을 JSON 형식으로 생성해주세요.
    
    주제: {topic}
    간략 내용: {description}
    총 강의 시간: {total_hours}시간
    
    다음 JSON 형식을 사용해주세요:
    {{
        "topic": "강의 주제",
        "description": "강의 설명",
        "total_hours": 총 강의 시간,
        "lectures": [
            {{
                "title": "강의 제목",
                "content": "강의 내용 (3-4줄)",
                "duration": 소요 시간(분)
            }},
            ...
        ]
    }}
    """
    
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "당신은 교육 전문가입니다. 주어진 주제에 대한 상세한 강의 커리큘럼을 JSON 형식으로 생성해주세요."},
            {"role": "user", "content": prompt}
        ]
    )
    
    # JSON 문자열을 파이썬 객체로 변환
    try:
        curriculum_json = json.loads(response.choices[0].message.content)
        return curriculum_json
    except json.JSONDecodeError as e:
        print(f"JSON 파싱 오류: {e}")
        return None

def save_to_excel(curriculum_json, output_file):
    """생성된 커리큘럼을 엑셀 파일로 저장"""
    if not curriculum_json:
        print("커리큘럼 데이터가 없습니다.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "강의 커리큘럼"
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # 강의 정보 헤더
    info_headers = ["항목", "내용"]
    for col, header in enumerate(info_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 강의 정보 작성
    ws.cell(row=2, column=1, value="강의 주제")
    ws.cell(row=2, column=2, value=curriculum_json["topic"])
    ws.cell(row=3, column=1, value="강의 설명")
    ws.cell(row=3, column=2, value=curriculum_json["description"])
    ws.cell(row=4, column=1, value="총 강의 시간")
    ws.cell(row=4, column=2, value=f"{curriculum_json['total_hours']}시간")
    
    # 강의 목록 헤더
    lecture_headers = ["강의 제목", "강의 내용", "소요 시간(분)"]
    for col, header in enumerate(lecture_headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 강의 목록 작성
    current_row = 7
    for lecture in curriculum_json["lectures"]:
        ws.cell(row=current_row, column=1, value=lecture["title"])
        ws.cell(row=current_row, column=2, value=lecture["content"])
        ws.cell(row=current_row, column=3, value=lecture["duration"])
        current_row += 1
    
    # 열 너비 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # 파일 저장
    wb.save(output_file)
    print(f"커리큘럼이 {output_file}에 저장되었습니다.")
    
    # JSON 파일도 함께 저장
    json_file = output_file.replace('.xlsx', '.json')
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(curriculum_json, f, ensure_ascii=False, indent=2)
    print(f"JSON 파일이 {json_file}에 저장되었습니다.")

def main():
    # 사용자 입력 받기
    # topic = input("강의 주제를 입력하세요: ")
    # description = input("강의 내용을 간단히 설명하세요: ")
    # total_hours = int(input("총 강의 시간(시간)을 입력하세요: "))
    
    topic = '생성형AI 기반 RPA'
    description = '인공지능의 새로운 트렌드인 생성형AI를 활용하여 다양한 업무자동화를 위한 소스코드를 생성하고, 필요한 멀티미디어를 생성하는 능력을 함양'
    total_hours = 12

    topic = 'CrewAI를 활용한 AI 멀티 에이전트 서비스 개발 프로젝트'
    description = 'AI 멀티 에이전트 서비스의 구조와 원리를 이해, CrewAI 기반의 AI 멀티 에이전트 서비스 개발 역량 향상'
    total_hours = 12

    # 현재 스크립트의 디렉토리 경로를 가져옴
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 커리큘럼 생성
    print("커리큘럼을 생성하는 중...")
    curriculum_json = generate_curriculum(topic, description, total_hours)
    
    if curriculum_json:
        # 엑셀 파일로 저장 (현재 스크립트 위치에 저장)
        output_file = os.path.join(current_dir, f"curriculum_{topic.replace(' ', '_')}.xlsx")
        save_to_excel(curriculum_json, output_file)
    else:
        print("커리큘럼 생성에 실패했습니다.")

if __name__ == "__main__":
    main()